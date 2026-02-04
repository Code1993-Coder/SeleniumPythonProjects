from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait
import openpyxl


#Finding the fruit cell number and price value of it
def update_excel_data(filepath,searchTerm,colName,new_price):
    file=openpyxl.load_workbook(filepath)
    sheet=file.active
    dict={}


    #get price column number
    for i in range(1,sheet.max_column+1):
        if sheet.cell(row=1,column=i).value==colName:
            dict['col']=i

    #get the row number of fruit

    for i in range(1,sheet.max_row+1):
        for j in range(1,sheet.max_column+1):
            if sheet.cell(row=i,column=j).value==searchTerm:
                dict['row']=i

    sheet.cell(row=dict['row'], column=dict['col']).value = new_price
    file.save(filepath)

print("***********************************************")
file_path="C:\\Users\\Manju Mohan\\Downloads\\download.xlsx"
fruit_name="Apple"
new_price="299"
driver=webdriver.Chrome()
driver.implicitly_wait(2)
driver.get("https://rahulshettyacademy.com/upload-download-test/")
driver.maximize_window()

#Dowload the excel
driver.find_element(By.CSS_SELECTOR,"#downloadButton").click()


#update the contents of the downloaded file
update_excel_data(file_path,fruit_name,"price",new_price)

#Upload the file-this only works if input type is file for sendkeys to work
driver.find_element(By.XPATH,"//input[@type='file']").send_keys(file_path)


#Wait for message to appear
wait=WebDriverWait(driver,5)
toast_locator=(By.CSS_SELECTOR,".Toastify__toast-body div:nth-child(2)")
wait.until(expected_conditions.visibility_of_element_located(toast_locator))
print(driver.find_element(*toast_locator).text)

#get me the price of the fruit
#Generic locator creation for getting the price for fruitname

#this get the dynamic position of price col
price_col=driver.find_element(By.XPATH,"//div[text()='Price']").get_attribute("data-column-id")

#this retrieves prices of fruit via price col
actual_price=driver.find_element(By.XPATH,"//div[text()='"+fruit_name+"']/parent::div/parent::div/div[@id='cell-"+price_col+"-undefined']").text
assert actual_price==new_price

