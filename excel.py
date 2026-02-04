#To work on execl-install openpyxl package via pip install openpyxl

import openpyxl


book=openpyxl.load_workbook(r"C:\Users\Manju Mohan\Desktop\Book1.xlsx")
sheet=book.active
dict={}
#to access row and col
cell=sheet.cell(row=1,column=1)

#to get value of the cell
print(cell.value)

#insert data to cell-this is not working
sheet.cell(row=2,column=2).value='Tina Williams'

#get the max row
print(sheet.max_row)

#get the max col
print(sheet.max_column)

#get sheet cell A2 value
print(sheet['A2'].value)
print("*****************************")


#iterating the table in excel
for r in range(1,sheet.max_row+1):#for rows
    if sheet.cell(row=r,column=1).value=='E001':
        for c in range(2,sheet.max_column+1):#for col
            dict[sheet.cell(row=1,column=c).value]=sheet.cell(row=r,column=c).value

print(dict)

